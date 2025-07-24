# evaluator/cell_locator.py

import re
import openpyxl


def build_activity_map(workbook,
                       sheet_name='EVALUACIÓN',
                       header_text="RESULTADO APRENDIZAJE",
                       activity_row=9):
    """
    Construye un mapa de actividades agrupado por trimestre.

    Retorna una lista de diccionarios, donde cada diccionario representa un trimestre.
    Ejemplo:
    [
        {'trimestre_name': '1er Trimestre', 'tasks': {'TAREA 1': 'D', 'TAREA 2': 'E'}},
        {'trimestre_name': '2do Trimestre', 'tasks': {'TAREA 1': 'L', 'TAREA 3': 'M'}},
    ]
    """
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

    # 1. Detectar los bloques de trimestres basados en las celdas combinadas del encabezado
    header_ranges = sorted(
        [
            r for r in sheet.merged_cells.ranges
            if header_text.lower() in str(sheet.cell(r.min_row, r.min_col).value or '').lower()
        ],
        key=lambda r: r.min_col  # Ordenar los trimestres de izquierda a derecha
    )

    if not header_ranges:
        raise ValueError(f"No se encontró ningún encabezado que contenga '{header_text}'.")

    # Patrón para encontrar "TAREA X" o "ACTIVIDAD X"
    pattern = re.compile(r"(TAREA|ACTIVIDAD)\s*(\d+)", re.IGNORECASE)

    trimester_map = []
    trimestre_labels = ["1er Trimestre", "2do Trimestre", "3er Trimestre"]

    # 2. Iterar sobre cada bloque de trimestre detectado
    for i, header_range in enumerate(header_ranges):

        current_trimester_tasks = {}
        # Recorrer solo las columnas que pertenecen a este bloque de trimestre
        for col_idx in range(header_range.min_col, header_range.max_col + 1):
            cell_value = sheet.cell(row=activity_row, column=col_idx).value
            if cell_value:
                match = pattern.search(str(cell_value))
                if match:
                    # Normalizar a "TAREA X"
                    task_name = f"TAREA {match.group(2)}"
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    # En este nuevo modelo, cada tarea solo tiene una columna por trimestre
                    current_trimester_tasks[task_name] = col_letter

        if current_trimester_tasks:
            trimestre_name = trimestre_labels[i] if i < len(trimestre_labels) else f"{i + 1}to Trimestre"
            trimester_map.append({
                'trimestre_name': trimestre_name,
                'tasks': current_trimester_tasks
            })

    if not trimester_map:
        raise ValueError("No se encontraron actividades con el formato 'TAREA X' en ningún trimestre.")

    return trimester_map