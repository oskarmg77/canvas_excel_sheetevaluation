# evaluator/mapping.py

import re
import openpyxl.utils
import logging


def _build_map_logic(sheet, header_ranges, activity_row):
    pattern = re.compile(r"(TAREA|ACTIVIDAD)\s*(\d+)", re.IGNORECASE)
    trimester_map = []
    trimestre_labels = ["1er Trimestre", "2do Trimestre", "3er Trimestre"]

    for i, header_range in enumerate(header_ranges):
        min_col = header_range.min_col if hasattr(header_range, 'min_col') else header_range['min_col']
        max_col = header_range.max_col if hasattr(header_range, 'max_col') else header_range['max_col']

        current_trimester_tasks = {}
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=activity_row, column=col_idx)
            if cell and cell.value:
                match = pattern.search(str(cell.value))
                if match:
                    task_name = f"TAREA {match.group(2)}"
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    current_trimester_tasks[task_name] = col_letter

        if current_trimester_tasks:
            trimestre_name = trimestre_labels[i] if i < len(trimestre_labels) else f"{i + 1}to Trimestre"
            trimester_map.append({
                'trimestre_name': trimestre_name,
                'tasks': current_trimester_tasks
            })

    if not trimester_map:
        raise ValueError("No se encontraron actividades con el formato 'TAREA X' en ningún trimestre.")

    logging.info(f"Mapa de actividades construido con éxito: {trimester_map}")
    return trimester_map


def build_map_from_excel(workbook, sheet_name='EVALUACIÓN', header_text="RESULTADO APRENDIZAJE", activity_row=9):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

    merged_ranges_raw = sheet.merged_cells.ranges if hasattr(sheet, 'merged_cells') else []
    header_ranges = sorted(
        [r for r in merged_ranges_raw if
         header_text.lower() in str(sheet.cell(r.min_row, r.min_col).value or '').lower()],
        key=lambda r: r.min_col
    )

    if not header_ranges:
        raise ValueError(f"No se encontró ningún encabezado que contenga '{header_text}'.")

    return _build_map_logic(sheet, header_ranges, activity_row)


def build_map_from_gsheet_data(data, header_text="RESULTADO APRENDIZAJE", activity_row=9):
    activity_row_index = activity_row - 1
    if len(data) <= activity_row_index:
        raise ValueError("Los datos de la hoja no tienen suficientes filas.")

    header_starts = []
    for row_to_check_idx in range(activity_row_index - 4, activity_row_index):
        if row_to_check_idx >= 0 and row_to_check_idx < len(data):
            for i, cell_value in enumerate(data[row_to_check_idx]):
                if header_text.lower() in str(cell_value).lower() and i not in header_starts:
                    header_starts.append(i)
    header_starts.sort()

    if not header_starts:
        raise ValueError(f"No se encontró el texto del encabezado '{header_text}'.")

    header_ranges = []
    for i, start_col in enumerate(header_starts):
        end_col = header_starts[i + 1] - 1 if i + 1 < len(header_starts) else len(data[activity_row_index]) - 1
        header_ranges.append({'min_col': start_col + 1, 'max_col': end_col + 1})

    class MockSheet:
        def cell(self, row, column):
            class MockCell:
                def __init__(self, value):
                    self.value = value

            try:
                return MockCell(data[row - 1][column - 1])
            except IndexError:
                return MockCell(None)

    return _build_map_logic(MockSheet(), header_ranges, activity_row)