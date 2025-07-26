# evaluator/processor.py

import logging
import json
import openpyxl
import os
import shutil
from datetime import datetime

import pandas as pd

# Importar los módulos necesarios del paquete 'evaluator'
from . import clients
from . import mapping
from . import matcher


# En evaluator/processor.py

def _write_grades_to_excel(file_path: str, grades_to_write: list) -> dict:
    """
    Motor de escritura para archivos Excel. Realiza el backup, busca coincidencias
    y escribe todas las notas.
    """
    logging.info(f"Iniciando proceso de escritura para Excel: {file_path}")

    # --- 1. Crear copia de seguridad ---
    try:
        root, ext = os.path.splitext(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{root}_backup_{timestamp}{ext}"
        shutil.copy2(file_path, backup_path)
        logging.info(f"Copia de seguridad creada en: {backup_path}")
    except Exception as e:
        logging.error(f"No se pudo crear la copia de seguridad: {e}", exc_info=True)
        raise IOError(f"No se pudo crear la copia de seguridad: {e}")

    # --- 2. Preparar el archivo y procesar ---
    # <<< CORRECCIÓN CLAVE: Abrir en modo data_only=True PARA LEER >>>
    workbook_read = openpyxl.load_workbook(file_path, data_only=True)
    sheet_read = workbook_read['EVALUACIÓN']

    written_count = 0
    not_found_students = []

    updates_to_perform = {}

    for record in grades_to_write:
        student_name = record.get('name')
        score = record.get('score')

        if not student_name or pd.isna(score):
            continue

        row_to_write = matcher.find_match_in_excel(sheet_read, student_name)

        if row_to_write:
            try:
                score_value = float(score)
                column = record['target_col']
                cell = f"{column}{row_to_write}"
                updates_to_perform[cell] = score_value
            except (ValueError, TypeError):
                logging.warning(f"Nota no válida para '{student_name}': {score}. Se omite.")
                continue
        else:
            not_found_students.append(student_name)
            logging.warning(f"No se encontró coincidencia para el alumno de Canvas: '{student_name}'")

    workbook_read.close()

    # --- 3. Escribir los cambios si hay algo que actualizar ---
    if updates_to_perform:
        logging.info(f"Escribiendo {len(updates_to_perform)} notas en el archivo Excel...")
        workbook_write = openpyxl.load_workbook(file_path)
        sheet_write = workbook_write['EVALUACIÓN']
        for cell, grade in updates_to_perform.items():
            sheet_write[cell].value = grade
        workbook_write.save(file_path)
        workbook_write.close()
        written_count = len(updates_to_perform)

    # --- 4. Devolver el resumen ---
    return {
        "processed": len(grades_to_write),
        "written": written_count,
        "not_found": len(not_found_students),
        "not_found_names": not_found_students,
        "backup_path": backup_path
    }


def _write_grades_to_gsheet(spreadsheet_id: str, trimester_map: list, grades_to_write: list) -> dict:
    """
    Motor de escritura para Google Sheets. Busca coincidencias y escribe todas las notas.
    """
    logging.info(f"Iniciando proceso de escritura para Google Sheet ID: {spreadsheet_id}")

    # --- 1. Leer los datos frescos de la hoja para buscar ---
    sheet_data = clients.get_gsheet_values(spreadsheet_id, "EVALUACIÓN!A1:Z50")

    written_count = 0
    not_found_students = []

    for record in grades_to_write:
        student_name = record.get('name')
        score = record.get('score')

        if not student_name or pd.isna(score):
            continue

        row_to_write = matcher.find_match_in_gsheet(sheet_data, student_name)

        if row_to_write:
            try:
                score_value = float(score)
                column = record['target_col']
                range_to_update = f"EVALUACIÓN!{column}{row_to_write}"
                clients.update_gsheet_values(spreadsheet_id, range_to_update, [[score_value]])
                written_count += 1
            except (ValueError, TypeError):
                logging.warning(f"Nota no válida para '{student_name}': {score}. Se omite.")
                continue
        else:
            not_found_students.append(student_name)
            logging.warning(f"No se encontró coincidencia para el alumno de Canvas: '{student_name}'")

    # --- 2. Devolver el resumen ---
    return {
        "processed": len(grades_to_write),
        "written": written_count,
        "not_found": len(not_found_students),
        "not_found_names": not_found_students,
        "backup_path": None  # No aplica para Google Sheets
    }


def run_grade_processing(dest_config: dict) -> dict:
    """
    Función principal que orquesta todo el proceso de cotejo y escritura.

    :param dest_config: Un diccionario con la configuración del destino.
                        Ej: {'type': 'excel', 'path': '...', 'trimestre': '...', 'tarea': '...'}
                        Ej: {'type': 'sheets', 'id': '...', 'trimestre': '...', 'tarea': '...'}
    :return: Un diccionario con el resumen de la operación.
    """
    logging.info(f"Procesador iniciado. Configuración de destino: {dest_config}")

    # --- 1. Cargar las notas de Canvas que se van a escribir ---
    try:
        with open('canvas_grades_to_write.json', 'r', encoding='utf-8') as f:
            grades_to_write = json.load(f)
        if not grades_to_write:
            raise ValueError("'canvas_grades_to_write.json' está vacío o no es válido.")
        logging.info(f"Cargadas {len(grades_to_write)} notas desde 'canvas_grades_to_write.json'.")
    except FileNotFoundError:
        raise FileNotFoundError(
            "El archivo 'canvas_grades_to_write.json' no existe. Asegúrate de seleccionar una tarea de Canvas primero.")

    # --- 2. Obtener la columna de destino a partir del mapa ---
    if dest_config['type'] == 'excel':
        workbook = openpyxl.load_workbook(dest_config['path'], data_only=True)
        trimester_map = mapping.build_map_from_excel(workbook)
    else:  # sheets
        sheet_data = clients.get_gsheet_values(dest_config['id'], "EVALUACIÓN!A1:Z50")
        trimester_map = mapping.build_map_from_gsheet_data(sheet_data)

    trimestre_info = next((t for t in trimester_map if t['trimestre_name'] == dest_config['trimestre']), None)
    target_column = trimestre_info['tasks'].get(dest_config['tarea']) if trimestre_info else None

    if not target_column:
        raise ValueError(
            f"No se pudo encontrar la columna para la tarea '{dest_config['tarea']}' en el trimestre '{dest_config['trimestre']}'.")

    # Añadir la columna de destino a cada registro para que las funciones de escritura la conozcan
    for record in grades_to_write:
        record['target_col'] = target_column

    # --- 3. Ejecutar el motor de escritura correspondiente y devolver el resultado ---
    if dest_config['type'] == 'excel':
        return _write_grades_to_excel(dest_config['path'], trimester_map, grades_to_write)
    else:  # sheets
        return _write_grades_to_gsheet(dest_config['id'], trimester_map, grades_to_write)