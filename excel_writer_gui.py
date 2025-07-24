# excel_writer_gui.py
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
# ¡Ahora la importación funcionará correctamente!
from evaluator.cell_locator import build_activity_map, find_student_row


class ExcelWriterGUI(tk.Tk):
    """
    Interfaz gráfica para probar la escritura de notas desde un CSV
    a un archivo Excel local.
    """

    def __init__(self):
        super().__init__()
        self.title("Probador de Escritura en Excel")
        self.geometry("600x600")

        # --- Variables de estado ---
        self.excel_path = tk.StringVar()
        self.csv_path = tk.StringVar()
        self.workbook = None
        self.activity_map = {}

        # --- Creación de Widgets ---
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill="both", expand=True)

        # -- Sección 1: Selección de Archivos --
        file_frame = ttk.LabelFrame(main_frame, text="Paso 1: Cargar Archivos", padding="10")
        file_frame.pack(fill="x", pady=5)

        ttk.Button(file_frame, text="Seleccionar Plantilla Excel (.xlsx)", command=self.load_excel).pack(fill="x",
                                                                                                         pady=5)
        ttk.Label(file_frame, textvariable=self.excel_path, wraplength=550).pack()

        ttk.Button(file_frame, text="Seleccionar Archivo de Notas (.csv)", command=self.load_csv).pack(fill="x", pady=5)
        ttk.Label(file_frame, textvariable=self.csv_path, wraplength=550).pack()

        # -- Sección 2: Mapeo y Acción --
        action_frame = ttk.LabelFrame(main_frame, text="Paso 2: Escribir Notas", padding="10")
        action_frame.pack(fill="x", pady=10)

        ttk.Label(action_frame, text="Asignar notas a la siguiente Tarea del Excel:").pack(anchor="w")
        self.combo_tareas = ttk.Combobox(action_frame, state="disabled")
        self.combo_tareas.pack(fill="x", pady=5)

        self.btn_escribir = ttk.Button(action_frame, text="Escribir Notas en Excel", command=self.write_grades_to_excel,
                                       state="disabled")
        self.btn_escribir.pack(pady=10, fill="x")

        # -- Sección 3: Registro de Actividad --
        log_frame = ttk.LabelFrame(main_frame, text="Registro de Actividad", padding="10")
        log_frame.pack(fill="both", expand=True, pady=5)

        self.log_text = tk.Text(log_frame, state="disabled", height=10, wrap="word", font=("Courier New", 9))
        self.log_text.pack(fill="both", expand=True)

    def log(self, message):
        """Añade un mensaje al cuadro de registro."""
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")

    def load_excel(self):
        path = filedialog.askopenfilename(title="Selecciona la plantilla Excel",
                                          filetypes=[("Archivos de Excel", "*.xlsx")])
        if not path: return

        self.excel_path.set(path)
        self.log(f"Plantilla Excel seleccionada: {path}")

        try:
            self.workbook = openpyxl.load_workbook(path)
            self.activity_map = build_activity_map(self.workbook)

            if not self.activity_map:
                messagebox.showwarning("Mapa Vacío",
                                       "No se encontraron tareas en el archivo Excel con el formato esperado.")
                return

            self.combo_tareas['values'] = list(self.activity_map.keys())
            self.combo_tareas.config(state="readonly")
            self.log("Mapa de tareas del Excel cargado correctamente.")
            self.check_if_ready()
        except Exception as e:
            messagebox.showerror("Error al leer Excel", f"No se pudo cargar o procesar el archivo Excel:\n{e}")
            self.log(f"Error: {e}")

    def load_csv(self):
        path = filedialog.askopenfilename(title="Selecciona el archivo de calificaciones",
                                          filetypes=[("Archivos CSV", "*.csv")])
        if not path: return

        self.csv_path.set(path)
        self.log(f"Archivo de notas seleccionado: {path}")
        self.check_if_ready()

    def check_if_ready(self):
        if self.excel_path.get() and self.csv_path.get():
            self.btn_escribir.config(state="normal")

    def write_grades_to_excel(self):
        target_task = self.combo_tareas.get()
        if not target_task:
            messagebox.showwarning("Tarea no seleccionada",
                                   "Por favor, selecciona una tarea del Excel para asignar las notas.")
            return

        try:
            df_calificaciones = pd.read_csv(self.csv_path.get())
            self.log(f"Leyendo {len(df_calificaciones)} notas desde el CSV.")

            sheet = self.workbook['EVALUACIÓN']

            # Asumimos que queremos escribir en la primera celda disponible para esa tarea
            target_cell_ref_start = self.activity_map[target_task][0]
            target_column_letter = re.match(r"([A-Z]+)", target_cell_ref_start).group(1)

            for _, row in df_calificaciones.iterrows():
                student_name = row['name']
                score = row.get('score')  # Usamos .get() por si la columna no existiera

                if pd.isna(score):
                    self.log(f"AVISO: Nota vacía para '{student_name}'. Se omite.")
                    continue

                student_row_index = find_student_row(sheet, student_name)

                if student_row_index:
                    target_cell = f"{target_column_letter}{student_row_index}"
                    sheet[target_cell].value = float(score)
                    self.log(f"OK: Escrita nota {score} para '{student_name}' en la celda {target_cell}.")
                else:
                    self.log(f"AVISO: Alumno '{student_name}' no encontrado en el Excel.")

            save_path = filedialog.asksaveasfilename(
                title="Guardar Excel modificado como...",
                defaultextension=".xlsx",
                filetypes=[("Archivos de Excel", "*.xlsx")],
                initialfile="calificaciones_actualizadas.xlsx"
            )

            if save_path:
                self.workbook.save(save_path)
                messagebox.showinfo("Proceso Completado",
                                    f"El archivo con las notas actualizadas se ha guardado en:\n{save_path}")
                self.log(f"--- Proceso completado. Archivo guardado en {save_path} ---")

        except Exception as e:
            messagebox.showerror("Error durante la escritura", f"Ocurrió un error:\n{e}")
            self.log(f"Error: {e}")


if __name__ == "__main__":
    app = ExcelWriterGUI()
    app.mainloop()