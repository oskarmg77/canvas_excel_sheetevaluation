# main_app.py

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
import re
import unicodedata
import csv
import os
import shutil
from datetime import datetime
import logging
from logging.handlers import RotatingFileHandler

# Importar los dos clientes
from evaluator import canvas_client
from evaluator.cell_locator import build_activity_map as build_map_from_excel
from evaluator import google_sheets_client

LOG_FILENAME = 'app.log'


def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    if logger.hasHandlers(): logger.handlers.clear()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh = RotatingFileHandler(LOG_FILENAME, maxBytes=1 * 1024 * 1024, backupCount=1)
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    logger.addHandler(ch)


setup_logging()


class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        logging.info("=====================================")
        logging.info("Iniciando la aplicación...")
        self.title("Automatización de Notas: Canvas a Excel/Sheets")
        self.geometry("1200x800")

        self.source_type = tk.StringVar(value="excel")
        self.excel_file_path = None
        self.workbook_readonly = None
        self.spreadsheet_id = None
        self.sheet_data = None

        self.df_canvas = None
        self.df_alumnos_del_curso = None
        self.cursos_canvas_dict = {}
        self.tareas_canvas_dict = {}
        self.trimester_data_map = []
        self._create_widgets()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10");
        main_frame.pack(fill="both", expand=True)
        top_frame = ttk.Frame(main_frame);
        top_frame.pack(fill="x", pady=5)
        canvas_frame = ttk.LabelFrame(top_frame, text="Paso 1: Seleccionar Origen (Canvas)", padding="10");
        canvas_frame.pack(side="left", fill="x", expand=True, padx=5)
        self.btn_conectar_canvas = ttk.Button(canvas_frame, text="Conectar y Cargar Cursos",
                                              command=self._load_canvas_courses);
        self.btn_conectar_canvas.pack(fill="x")
        ttk.Label(canvas_frame, text="Curso de Canvas:").pack(anchor="w", pady=(5, 0))
        self.combo_canvas_cursos = ttk.Combobox(canvas_frame, state="disabled");
        self.combo_canvas_cursos.pack(fill="x")
        self.combo_canvas_cursos.bind("<<ComboboxSelected>>", self._on_course_selected)
        ttk.Label(canvas_frame, text="Tarea de Canvas:").pack(anchor="w", pady=(5, 0))
        self.combo_canvas_tareas = ttk.Combobox(canvas_frame, state="disabled");
        self.combo_canvas_tareas.pack(fill="x")
        self.combo_canvas_tareas.bind("<<ComboboxSelected>>", self._fetch_and_display_canvas_data)

        dest_frame = ttk.LabelFrame(top_frame, text="Paso 2: Seleccionar Destino", padding="10");
        dest_frame.pack(side="left", fill="x", expand=True, padx=5)
        source_chooser_frame = ttk.Frame(dest_frame);
        source_chooser_frame.pack(fill="x", pady=(0, 10))
        ttk.Radiobutton(source_chooser_frame, text="Excel Local", variable=self.source_type, value="excel",
                        command=self._on_source_change).pack(side="left", padx=5)
        ttk.Radiobutton(source_chooser_frame, text="Google Sheets", variable=self.source_type, value="sheets",
                        command=self._on_source_change).pack(side="left", padx=5)

        self.excel_controls_frame = ttk.Frame(dest_frame)
        excel_button_frame = ttk.Frame(self.excel_controls_frame);
        excel_button_frame.pack(fill="x")
        self.btn_load_excel = ttk.Button(excel_button_frame, text="Seleccionar Plantilla Excel",
                                         command=self._select_excel_file);
        self.btn_load_excel.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.btn_refresh_excel = ttk.Button(excel_button_frame, text="Refrescar Excel",
                                            command=self._on_refresh_button_pressed, state="disabled");
        self.btn_refresh_excel.pack(side="left", fill="x", expand=True)

        self.sheets_controls_frame = ttk.Frame(dest_frame)
        ttk.Label(self.sheets_controls_frame, text="URL de la Hoja de Google:").pack(anchor="w")
        self.entry_gsheet_url = ttk.Entry(self.sheets_controls_frame);
        self.entry_gsheet_url.pack(fill="x", pady=(0, 5))
        self.btn_load_gsheet = ttk.Button(self.sheets_controls_frame, text="Cargar Hoja de Google",
                                          command=self._load_google_sheet);
        self.btn_load_gsheet.pack(fill="x")

        ttk.Label(dest_frame, text="Trimestre de Destino:").pack(anchor="w", pady=(5, 0))
        self.combo_trimestre = ttk.Combobox(dest_frame, state="disabled");
        self.combo_trimestre.pack(fill="x")
        self.combo_trimestre.bind("<<ComboboxSelected>>", self._on_trimestre_selected)
        ttk.Label(dest_frame, text="Tarea de Destino:").pack(anchor="w", pady=(5, 0))
        self.combo_excel_tareas = ttk.Combobox(dest_frame, state="disabled");
        self.combo_excel_tareas.pack(fill="x")
        self.combo_excel_tareas.bind("<<ComboboxSelected>>", self._on_task_selected)

        viz_frame = ttk.LabelFrame(main_frame, text="Paso 3: Comparar Datos", padding="10");
        viz_frame.pack(fill="both", expand=True, pady=10)
        tables_container = ttk.Frame(viz_frame);
        tables_container.pack(fill="both", expand=True)
        canvas_table_frame = ttk.Frame(tables_container);
        canvas_table_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        ttk.Label(canvas_table_frame, text="Datos de Canvas", font=("", 10, "bold")).pack()
        self.tree_canvas = self._create_treeview(canvas_table_frame, ("Alumno Canvas", "Nota"))
        excel_table_frame = ttk.Frame(tables_container);
        excel_table_frame.pack(side="left", fill="both", expand=True, padx=(5, 0))
        ttk.Label(excel_table_frame, text="Datos de Destino", font=("", 10, "bold")).pack()
        self.tree_excel = self._create_treeview(excel_table_frame, ("Alumno Destino", "Nota Actual"))
        action_frame = ttk.Frame(main_frame);
        action_frame.pack(fill="x", pady=5)
        self.btn_escribir = ttk.Button(action_frame, text="Generar Reporte y Escribir Notas",
                                       command=self._execute_write_process, state="disabled");
        self.btn_escribir.pack(side="left", fill="x", expand=True, ipady=5, padx=5)
        self.btn_log = ttk.Button(action_frame, text="Ver Registro de Actividad", command=self._show_log_window);
        self.btn_log.pack(side="right", ipady=5, padx=5)

        self._on_source_change()

    def _on_source_change(self):
        source = self.source_type.get()
        logging.info(f"Cambiando a modo de destino: {source.upper()}")
        if source == "excel":
            self.sheets_controls_frame.pack_forget()
            self.excel_controls_frame.pack(fill="x")
        else:  # sheets
            self.excel_controls_frame.pack_forget()
            self.sheets_controls_frame.pack(fill="x")
        self._clear_dest_data()
        self._check_if_ready_to_write()

    def _clear_dest_data(self):
        self.trimester_data_map = []
        self.combo_trimestre.set('')
        self.combo_trimestre.config(state="disabled")
        self.combo_excel_tareas.set('')
        self.combo_excel_tareas.config(state="disabled")
        self._clear_treeview(self.tree_excel)

    def _get_spreadsheet_id_from_url(self, url):
        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
        return match.group(1) if match else None

    def _load_google_sheet(self):
        url = self.entry_gsheet_url.get().strip()
        if not url: messagebox.showwarning("URL Vacía", "Por favor, pega la URL de tu Google Sheet."); return
        spreadsheet_id = self._get_spreadsheet_id_from_url(url)
        if not spreadsheet_id: messagebox.showerror("URL Inválida",
                                                    "La URL no parece ser válida para una hoja de Google."); return
        self.spreadsheet_id = spreadsheet_id
        logging.info(f"Cargando datos de Google Sheet con ID: {self.spreadsheet_id}")
        try:
            self.sheet_data = google_sheets_client.get_values(self.spreadsheet_id, "EVALUACIÓN!A1:Z50")
            if not self.sheet_data:
                raise ValueError("La hoja de Google parece estar vacía o no se pudieron leer datos.")
            self.trimester_data_map = google_sheets_client.build_activity_map_from_api_data(self.sheet_data)
            trimestre_names = [t['trimestre_name'] for t in self.trimester_data_map]
            self.combo_trimestre['values'] = trimestre_names
            if trimestre_names:
                self.combo_trimestre.config(state="readonly");
                self.combo_trimestre.set(trimestre_names[0])
            self._on_trimestre_selected()
            logging.info("Hoja de Google cargada y mapeada correctamente.")
            messagebox.showinfo("Éxito", "Hoja de Google cargada y mapeada correctamente.")
        except Exception as e:
            logging.error(f"Fallo al cargar la hoja de Google: {e}", exc_info=True)
            messagebox.showerror("Error al Cargar", f"No se pudo cargar o procesar la hoja de Google:\n{e}")
            self.spreadsheet_id = None

            def _execute_write_process(self):
                logging.info("Iniciando el proceso de generación de reporte y escritura.")
                trimestre_str = self.combo_trimestre.get()
                target_task = self.combo_excel_tareas.get()
                if not all([self.df_canvas is not None, trimestre_str, target_task]):
                    messagebox.showwarning("Faltan datos", "Asegúrate de haber seleccionado todos los campos.");
                    return

                trimestre_info = next((t for t in self.trimester_data_map if t['trimestre_name'] == trimestre_str),
                                      None)
                target_col = trimestre_info['tasks'].get(target_task) if trimestre_info else None
                if not target_col: messagebox.showerror("Error Interno",
                                                        "No se encontró la columna para la tarea/trimestre seleccionado."); return

                report_data = []
                for _, row in self.df_canvas.iterrows():
                    name_canvas, score = row.get('name'), row.get('score')
                    row_idx, name_dest = (None, None)
                    if self.source_type.get() == 'excel':
                        sheet = self.workbook_readonly['EVALUACIÓN']
                        row_idx, name_dest = self._find_student_row_excel(sheet, name_canvas)
                    else:
                        row_idx, name_dest = self._find_student_row_gsheet(self.sheet_data, name_canvas)
                    report_data.append({
                        "canvas_name": name_canvas, "score": score if not pd.isna(score) else 'N/A',
                        "excel_name": name_dest or '---',
                        "status": "Coincidencia" if row_idx else "NO ENCONTRADO", "excel_row": row_idx or 'N/A',
                        "target_col": target_col
                    })
                self._show_matching_report(report_data)

            def _show_matching_report(self, report_data):
                report_window = tk.Toplevel(self)
                report_window.title("Reporte de Cotejo")
                frame = ttk.Frame(report_window, padding="10");
                frame.pack(fill="both", expand=True)
                cols = ("Nombre Canvas", "Nombre Destino", "Estado", "Fila", "Columna")
                tree = self._create_treeview(frame, cols)
                for col in cols: tree.column(col, width=250, anchor="w")
                tree.tag_configure('found', foreground='green');
                tree.tag_configure('notfound', foreground='red')
                for item in report_data:
                    tags = ('found',) if item["status"] == "Coincidencia" else ('notfound',)
                    values = (item["canvas_name"], item["excel_name"], item["status"], item["excel_row"],
                              item["target_col"])
                    tree.insert("", "end", values=values, tags=tags)
                button_frame = ttk.Frame(report_window, padding="10");
                button_frame.pack(fill="x")

                def _proceed_to_write():
                    coincidencias = [item for item in report_data if item["status"] == "Coincidencia"]
                    if not coincidencias: messagebox.showwarning("Sin Coincidencias", "No hay notas que escribir.",
                                                                 parent=report_window); return
                    if self.source_type.get() == 'excel':
                        self._write_to_excel(coincidencias, report_window)
                    else:  # sheets
                        self._write_to_gsheet(coincidencias, report_window)

                btn_write = ttk.Button(button_frame, text="Confirmar y Escribir Notas", command=_proceed_to_write)
                btn_write.pack(side="left", fill="x", expand=True, padx=5)

            def _write_to_excel(self, data, parent_window):
                confirm = messagebox.askyesno("Confirmar Sobrescritura",
                                              "Estás a punto de sobrescribir el archivo Excel original.\n\nSe creará una copia de seguridad.\n\n¿Deseas continuar?",
                                              parent=parent_window)
                if not confirm: logging.info("El usuario canceló la operación de escritura."); return
                backup_path = ""
                try:
                    root, ext = os.path.splitext(self.excel_file_path)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_path = f"{root}_backup_{timestamp}{ext}"
                    shutil.copy2(self.excel_file_path, backup_path)
                    workbook_to_write = openpyxl.load_workbook(self.excel_file_path)
                    sheet_to_write = workbook_to_write['EVALUACIÓN']
                    written_count = 0
                    for item in data:
                        try:
                            score_value = float(item["score"])
                            sheet_to_write[f"{item['target_col']}{item['excel_row']}"].value = score_value
                            written_count += 1
                        except (ValueError, TypeError):
                            continue
                    workbook_to_write.save(self.excel_file_path)
                    workbook_to_write.close()
                    parent_window.destroy()
                    messagebox.showinfo("Proceso Completado",
                                        f"{written_count} notas escritas en:\n{os.path.basename(self.excel_file_path)}\n\nCopia de seguridad creada en:\n{os.path.basename(backup_path)}",
                                        parent=self)
                    self._refresh_excel_data(self.excel_file_path)
                except Exception as e:
                    messagebox.showerror("Error Durante la Escritura", f"Ocurrió un error: {e}", parent=parent_window)

            def _write_to_gsheet(self, data, parent_window):
                logging.info(f"Iniciando escritura de {len(data)} notas en Google Sheets.")
                written_count = 0
                try:
                    updates = []
                    for item in data:
                        try:
                            score_value = float(item["score"])
                            range_to_update = f"EVALUACIÓN!{item['target_col']}{item['excel_row']}"
                            # Esto se puede optimizar en el futuro para una sola llamada a la API (batchUpdate)
                            google_sheets_client.update_values(self.spreadsheet_id, range_to_update, [[score_value]])
                            written_count += 1
                        except (ValueError, TypeError):
                            continue
                    parent_window.destroy()
                    messagebox.showinfo("Proceso Completado",
                                        f"{written_count} notas han sido escritas con éxito en la hoja de Google.",
                                        parent=self)
                    self._load_google_sheet()
                except Exception as e:
                    logging.error(f"Error al escribir en Google Sheets: {e}", exc_info=True)
                    messagebox.showerror("Error de Escritura", f"No se pudo escribir en Google Sheets:\n{e}",
                                         parent=parent_window)

            def _normalize_name(self, text):
                if not isinstance(text, str): return ""
                stop_words = {'de', 'la', 'del', 'los', 'las', 'y', 'e', 'maria'}
                text = text.lower().replace(',', '')
                nfkd_form = unicodedata.normalize('NFD', text)
                text_no_accents = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
                words = text_no_accents.split()
                return " ".join([word for word in words if word not in stop_words])

            def _find_student_row_excel(self, sheet, name_canvas, col='C', start=10):
                canvas_parts = set(self._normalize_name(name_canvas).split())
                if not canvas_parts: return None, None
                best_match = {'row': None, 'name': None, 'score': -1}
                for row_idx in range(start, 45):
                    cell_value = sheet[f"{col}{row_idx}"].value
                    if not cell_value: continue
                    excel_parts = set(self._normalize_name(str(cell_value)).split())
                    common_words = len(canvas_parts.intersection(excel_parts))
                    if common_words > best_match['score']:
                        best_match = {'row': row_idx, 'name': cell_value, 'score': common_words}
                is_match = (best_match['score'] >= 2 or (len(canvas_parts) <= 2 and best_match['score'] >= 1))
                return (best_match['row'], best_match['name']) if is_match else (None, None)

            def _find_student_row_gsheet(self, data, name_canvas, col_idx=2, start=10):
                canvas_parts = set(self._normalize_name(name_canvas).split())
                if not canvas_parts: return None, None
                best_match = {'row': None, 'name': None, 'score': -1}
                for i, row_data in enumerate(data[start - 1:44]):  # Limitar a la fila 44
                    row_idx = i + start
                    if len(row_data) > col_idx and row_data[col_idx]:
                        cell_value = row_data[col_idx]
                        excel_parts = set(self._normalize_name(cell_value).split())
                        common_words = len(canvas_parts.intersection(excel_parts))
                        if common_words > best_match['score']:
                            best_match = {'row': row_idx, 'name': cell_value, 'score': common_words}
                is_match = (best_match['score'] >= 2 or (len(canvas_parts) <= 2 and best_match['score'] >= 1))
                return (best_match['row'], best_match['name']) if is_match else (None, None)

            def _load_canvas_courses(self):
                try:
                    self.cursos_canvas_dict = canvas_client.obtener_cursos()
                    self.combo_canvas_cursos['values'] = list(self.cursos_canvas_dict.keys())
                    self.combo_canvas_cursos.config(state="readonly")
                except Exception as e:
                    messagebox.showerror("Error de Conexión", f"No se pudo conectar a Canvas: {e}")

            def _on_course_selected(self, event=None):
                nombre_curso = self.combo_canvas_cursos.get()
                if not nombre_curso: return
                curso_id = self.cursos_canvas_dict[nombre_curso]
                try:
                    self.tareas_canvas_dict = canvas_client.obtener_tareas(curso_id)
                    self.combo_canvas_tareas['values'] = list(self.tareas_canvas_dict.keys())
                    self.combo_canvas_tareas.config(state="readonly")
                    self.df_alumnos_del_curso = canvas_client.obtener_alumnos(curso_id)
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudieron cargar los datos del curso: {e}")

            def _fetch_and_display_canvas_data(self, event=None):
                nombre_tarea = self.combo_canvas_tareas.get()
                if not nombre_tarea or self.df_alumnos_del_curso is None: return
                curso_id = self.cursos_canvas_dict[self.combo_canvas_cursos.get()]
                tarea_id = self.tareas_canvas_dict[nombre_tarea]
                try:
                    df_calificaciones = canvas_client.obtener_calificaciones(curso_id, tarea_id)
                    df_alumnos_renamed = self.df_alumnos_del_curso.rename(columns={'id': 'user_id'})
                    self.df_canvas = pd.merge(df_calificaciones, df_alumnos_renamed[['user_id', 'name']], on='user_id',
                                              how='right')
                    self.df_canvas.dropna(subset=['name'], inplace=True)

                    def round_score(score):
                        try:
                            return round(float(score), 1)
                        except (ValueError, TypeError):
                            return score

                    self.df_canvas['score'] = self.df_canvas['score'].apply(round_score)
                    self._clear_treeview(self.tree_canvas)
                    for _, row in self.df_canvas.iterrows():
                        self.tree_canvas.insert("", "end", values=(row.get('name', 'N/A'), row.get('score', 'N/A')))
                    self._check_if_ready_to_write()
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudieron obtener los datos de Canvas: {e}")

            def _select_excel_file(self):
                path = filedialog.askopenfilename(title="Selecciona la plantilla Excel",
                                                  filetypes=[("Archivos de Excel", "*.xlsx")])
                if path: self._refresh_excel_data(path)

            def _refresh_excel_data(self, path):
                try:
                    self.excel_file_path = path
                    self.workbook_readonly = openpyxl.load_workbook(path, data_only=True)
                    self.trimester_data_map = build_map_from_excel(self.workbook_readonly)
                    sheet = self.workbook_readonly['EVALUACIÓN']
                    alumnos_encontrados = any(sheet[f'C{row_idx}'].value for row_idx in range(10, 45))
                    if not alumnos_encontrados:
                        messagebox.showerror("Error de Lectura de Excel",
                                             "No se pudieron leer los nombres de los alumnos del archivo.\n\nCausa probable: El archivo fue guardado por un programa que no es Excel.\n\nSolución: Abra el archivo en Excel, guárdelo y vuelva a cargarlo aquí.")
                        return
                    trimestre_names = [t['trimestre_name'] for t in self.trimester_data_map]
                    self.combo_trimestre['values'] = trimestre_names
                    if trimestre_names: self.combo_trimestre.set(trimestre_names[0])
                    self._on_trimestre_selected()
                    self.btn_refresh_excel.config(state="normal")
                except Exception as e:
                    messagebox.showerror("Error al leer Excel", f"No se pudo procesar el archivo Excel:\n{e}")

            def _on_trimestre_selected(self, event=None):
                trimestre_str = self.combo_trimestre.get()
                if not trimestre_str: return
                trimestre_info = next((t for t in self.trimester_data_map if t['trimestre_name'] == trimestre_str),
                                      None)
                if trimestre_info and trimestre_info['tasks']:
                    task_names = sorted(list(trimestre_info['tasks'].keys()))
                    self.combo_excel_tareas['values'] = task_names
                    self.combo_excel_tareas.set(task_names[0])
                else:
                    self.combo_excel_tareas.set('')
                self._on_task_selected()

            def _on_task_selected(self, event=None):
                self._clear_treeview(self.tree_excel)
                source = self.source_type.get()
                if source == 'excel' and self.excel_file_path:
                    self._display_excel_data()
                elif source == 'sheets' and self.sheet_data:
                    self._display_gsheet_data()
                self._check_if_ready_to_write()

            def _display_excel_data(self):
                trimestre_str = self.combo_trimestre.get()
                target_task = self.combo_excel_tareas.get()
                if not all([trimestre_str, target_task, self.workbook_readonly]): return
                trimestre_info = next((t for t in self.trimester_data_map if t['trimestre_name'] == trimestre_str),
                                      None)
                target_col_letter = trimestre_info['tasks'].get(target_task) if trimestre_info else None
                if not target_col_letter: return
                try:
                    target_col_index = openpyxl.utils.column_index_from_string(target_col_letter)
                    sheet = self.workbook_readonly['EVALUACIÓN']
                    excel_students_data = []
                    for row_idx in range(10, 45):
                        student_name = sheet[f'C{row_idx}'].value
                        if not student_name: continue
                        grade = sheet.cell(row=row_idx, column=target_col_index).value
                        excel_students_data.append({'name': student_name, 'grade': grade if grade is not None else ''})
                    excel_students_data.sort(key=lambda x: str(x['name']).lower())
                    for student_data in excel_students_data:
                        self.tree_excel.insert("", "end", values=(student_data['name'], student_data['grade']))
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudieron leer los datos de la tarea de Excel: {e}")

            def _display_gsheet_data(self):
                trimestre_str = self.combo_trimestre.get()
                target_task = self.combo_excel_tareas.get()
                if not all([trimestre_str, target_task, self.sheet_data]): return
                trimestre_info = next((t for t in self.trimester_data_map if t['trimestre_name'] == trimestre_str),
                                      None)
                target_col_letter = trimestre_info['tasks'].get(target_task) if trimestre_info else None
                if not target_col_letter: return
                try:
                    target_col_index = openpyxl.utils.column_index_from_string(target_col_letter) - 1
                    gsheet_students_data = []
                    for row_idx in range(9, 44):
                        if row_idx < len(self.sheet_data):
                            row_data = self.sheet_data[row_idx]
                            if len(row_data) > 2:
                                student_name = row_data[2]
                                grade = row_data[target_col_index] if len(row_data) > target_col_index else ''
                                if student_name:  # Solo añadir si hay nombre de alumno
                                    gsheet_students_data.append({'name': student_name, 'grade': grade})
                    gsheet_students_data.sort(key=lambda x: str(x['name']).lower())
                    for student_data in gsheet_students_data:
                        self.tree_excel.insert("", "end", values=(student_data['name'], student_data['grade']))
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudieron leer los datos de Google Sheets: {e}")

            def _on_refresh_button_pressed(self):
                if self.excel_file_path:
                    self._refresh_excel_data(self.excel_file_path)
                    self.after(100, lambda: messagebox.showinfo("Refresco completado",
                                                                f"Se han recargado los datos del archivo:\n{os.path.basename(self.excel_file_path)}",
                                                                parent=self))

            def _check_if_ready_to_write(self):
                source_ready = (self.source_type.get() == 'excel' and self.excel_file_path) or \
                               (self.source_type.get() == 'sheets' and self.spreadsheet_id)
                if all([self.df_canvas is not None, self.combo_excel_tareas.get(), self.combo_trimestre.get(),
                        source_ready]):
                    self.btn_escribir.config(state="normal")
                else:
                    self.btn_escribir.config(state="disabled")

            def _show_log_window(self):
                logging.info("Abriendo ventana de registro de actividad.")
                log_window = tk.Toplevel(self)
                log_window.title("Registro de Actividad")
                log_window.geometry("800x600")
                text_area = tk.Text(log_window, wrap="word", state="normal", font=("Courier New", 9))
                vsb = ttk.Scrollbar(log_window, orient="vertical", command=text_area.yview)
                text_area.configure(yscrollcommand=vsb.set)
                vsb.pack(side="right", fill="y");
                text_area.pack(fill="both", expand=True)
                try:
                    with open(LOG_FILENAME, 'r', encoding='utf-8') as f:
                        log_content = f.read()
                    text_area.insert("1.0", log_content)
                except Exception as e:
                    text_area.insert("1.0", f"No se pudo leer el archivo de registro: {e}")
                text_area.config(state="disabled");
                text_area.see(tk.END)

        if __name__ == "__main__":
            app = MainApp()
            app.mainloop()